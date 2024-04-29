# IMPORTS
import openpyxl
import numpy as np
import time
import logging
import os
import yaml
import pandas as pd
import factory_flexibility_model.simulation.Simulation as fs
import dri_functions.dri_factory_definitions as dri
from multiprocessing import Pool
from factory_flexibility_model.io.set_logger import set_logging_level

# FUNCTIONS
def simulate_dri(simulation_task, model_parameters, simulation_config, total_sim_count):
    simulation_filename = f"{simulation_task['layout']}_{simulation_task['timeseries']}_NG{simulation_task['natural_gas_cost']}_EL{simulation_task['avg_electricity_price']}_VOL{simulation_task['volatility']}_CO2{simulation_task['co2_price']}_Reduction{simulation_task['co2_reduction']}_{simulation_task['month']}_ElectricityEmissions_{simulation_task['electricity_emissions']}"



    if not simulation_config["overwrite_existing_simulations"]:
        # check if simulation is already solved
        if os.path.isfile(rf"{simulation_config['filepath_solved']}\{simulation_filename}.sim") or os.path.isfile(rf"{simulation_config['filepath_problem']}\{simulation_filename}.sim"):
            # if simulation_config["enable_log_simulation_iteration"]:
            #     print(f"{round(simulation_task['simulation_number'] / total_sim_count * 100, 2)}%     Simulation {simulation_task['simulation_number']} is already solved")
            return

    # get scenario
    scenario = simulation_task["scenario"]
    t_start = time.time()

    # set prices
    scenario.configurations[simulation_task["factory"].get_key("Source Natural Gas")] = {"cost": simulation_task["natural_gas_cost"]}

    # scale timeseries to scenario parameters -> a explanation of this method can be found in the papers appendix
    cost_electricity = simulation_task["cost_electricity"]
    timeseries_scaling_factor = simulation_task["avg_electricity_price"] * simulation_task["volatility"] / np.std(cost_electricity)
    cost_electricity = (cost_electricity - cost_electricity.mean()) * timeseries_scaling_factor + simulation_task["avg_electricity_price"]
    scenario.configurations[simulation_task["factory"].get_key("Electricity Grid")] = {"cost": cost_electricity,
                                                                                       "co2_emissions_per_unit": simulation_task["emissions_electricity"]}

    # set slack costs
    scenario.configurations[simulation_task["factory"].get_key("Electricity Slack")] = {"cost": 1000000,
                                                                                        "co2_emissions_per_unit": 0}
    scenario.configurations[simulation_task["factory"].get_key("CO2 Slack")] = {"cost": 1000000,
                                                                                "co2_emissions_per_unit": 0}
    scenario.configurations[simulation_task["factory"].get_key("CO2 Emissions")] = {"co2_emissions_per_unit": 1}

    # set emission limit
    emission_limit = model_parameters["annual_dri_production_mtons"] \
                     * 1000000 * model_parameters["emission_baseline"] \
                     * (1 - simulation_task["co2_reduction"] / 100) \
                     * simulation_task["factory"].timesteps / 8760
    simulation_task["factory"].emission_accounting = True
    simulation_task["factory"].emission_limit = emission_limit
    simulation_task["factory"].emission_cost = simulation_task["co2_price"]

    # perform simulation
    if simulation_config["enable_log_simulation_setup"]:
        logging.getLogger().setLevel(logging.DEBUG)
    else:
        logging.getLogger().setLevel(logging.ERROR)
    simulation = fs.Simulation(factory=simulation_task["factory"],
                               scenario=scenario)
    simulation.simulate(threshold=simulation_config["threshold"],
                        solver_config={"max_solver_time": simulation_config["max_solver_time"],
                                       "mip_gap": simulation_config["mip_gap"],
                                       "log_solver": simulation_config["enable_log_solver"]})

    # write metadata into simulation object
    simulation_task["solver_time"] = time.time()-t_start
    simulation.info = simulation_task


    if simulation.simulated:
        simulation.save(simulation_config["filepath_solved"],
                        name=simulation_filename,
                        overwrite=True)
        if simulation_config["enable_log_simulation_iteration"]:
            solved_sim_count = len(os.listdir(simulation_config['filepath_solved'])) + len(
                os.listdir(simulation_config['filepath_problem']))
            print(f"({solved_sim_count}/{total_sim_count} | {round(solved_sim_count/total_sim_count*100, 2)}%) \t Simulation number {simulation_task['simulation_number']} finished within: {round((time.time()-t_start), 2)} s")
    else:
        simulation.save(simulation_config["filepath_problem"],
                        name=simulation_filename,
                        overwrite=True)
        if simulation_config["enable_log_simulation_iteration"]:
            solved_sim_count = len(os.listdir(simulation_config['filepath_solved'])) + len(
                os.listdir(simulation_config['filepath_problem']))
            print(f"({solved_sim_count}/{total_sim_count} | {round(solved_sim_count/total_sim_count*100, 2)}%) \t Simulation number {simulation_task['simulation_number']} aborted due to solver timeout")


def simulate_dri_wrapper(simulation_task, model_parameters, simulation_config, total_sim_count):
    """
    This function acts just as a utility for using the parallel pool. It takes the arguments for a single simulation execution out of the args_list and calls the simulate_dri function while correctly assigning the positional arguments
    """
    return simulate_dri(simulation_task, model_parameters,  simulation_config=simulation_config, total_sim_count=total_sim_count)


def worker_init():
    logging.getLogger().setLevel(logging.DEBUG)


def iterate_dri_simulations():
    """
    DOCSTRING SCHREIBEN!!!
    """

    # import input data
    simulation_config, model_parameters, timeseries_data, scenario_variations = import_input_data()

    # create a list with all required runs
    simulation_list = create_simulation_list(simulation_config, model_parameters, timeseries_data, scenario_variations)

    # make sure that the required output folders exist
    #simulation_config["filepath_solved"] = "F:\\FactoryFlexibilityModel\\DRI Simulations\\simulations_solved"
    #simulation_config["filepath_problem"] = "F:\\FactoryFlexibilityModel\\DRI Simulations\\simulations_failed"
    if not os.path.exists(simulation_config["filepath_solved"]):
        os.makedirs(simulation_config["filepath_solved"])
    if not os.path.exists(simulation_config["filepath_problem"]):
        os.makedirs(simulation_config["filepath_problem"])


    # iterate over the list of required simulations and call the simulation routine

    if simulation_config["parallel_workers"] == 1:
        # if only one worker is required: just iterate over the list

        for simulation_task in simulation_list:
            simulate_dri(simulation_task,
                         model_parameters,
                         simulation_config=simulation_config,
                         total_sim_count=count_simulations(scenario_variations))
    else:
        # solve simulations using pool processing
        args_list = [(simulation_task, model_parameters, simulation_config, count_simulations(scenario_variations))
                     for simulation_task in simulation_list]

        with Pool(processes=simulation_config["parallel_workers"], initializer=worker_init) as pool:
            pool.starmap(simulate_dri_wrapper, args_list)


def import_input_data():
    """
    This function imports all the required input parameters for the simulations:

    1) simulation_config: A config file containing logging- and solver preferences:
        - max_solver_time: [int] -> seconds till solver abort per run
        - enable_log_factory_setup: [boolean] -> enables logging while creating the plantlayout-objects
        - enable_log_simulation_setup: [boolean] -> enables logging while creating the optimization problem
        - enable_log_solver: [boolean] -> enables logging during solver runtime
        - enable_log_simulation_iteration: [boolean -> enables logging of each runs solver time in the console
        - enable_slacks: [boolean] -> adds slacks to all components of the plant layout to test for infeasability-causes
        - overwrite_existing_simulations: [boolean] -> Set to True if existing simulations runs shall be resolved and overwritten. Otherwise they are skipped
        - threshold: lower threshold for values to be rounded to 0 during result processing

        Source: "simulations\\input_data\\simulation_config.txt"


    2) model_parameters: List of materialbalance and process parameters that defines the exact process characteristics of the dri plants
        Full list of input parameters, assumptions and literature sources can be found in the excel file

        Source: "simulations\\input_data\\DRI Parameters.xlsx" | Sheet: "1 Model Parameters"


    3) timeseries_data: Collection of all required timeseries input data
        Contains all timeseries specified columnwise in the excelsheet

        Source: "simulations\\input_data\\DRI Parameters.xlsx" | Sheet: "2 Market Timeseries"


    4) scenario_variations: Lists of parameter variations for every variable simulation parameter:
        - plant_types [string-keys]
        - CO2_prices [float]
        - avg_electricity_cost [float]
        - market_timeseries [string-keys]
        - months [int 1..12]
        - natural_gas [float]
        - volatilities [float]

        Source: "simulations\\input_data\\DRI Parameters.xlsx" | Sheet: "3 Scenario Variations"


    :return: simulation_config [dict]; model_parameters [dict]; timeseries_data [pd.dataframe], scenario_variations [pd.dataframe]
    """

    # 1) import config
    with open(f"{os.getcwd()}\\simulations\\simulation_config.txt", "r") as file:
        simulation_config = yaml.load(file, Loader=yaml.SafeLoader)

    # 2) import model parameters
    config_path = f"{os.getcwd()}\\simulations\\input_data\\DRI Parameters.xlsx"
    wb = openpyxl.load_workbook(config_path, data_only=True)
    ws = wb["1 Model Parameters"]
    model_parameters = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        key = row[1]
        value = row[2]
        if key:
            model_parameters[key] = value

    # 3) import timeseries
    timeseries_data = pd.read_excel(config_path, sheet_name="2 Market Timeseries")

    # 4) import scenario_variations
    ws = wb["3 Scenario Variations"]
    scenario_variations = {}
    for column in ws.iter_cols(values_only=True):
        key = column[0]
        if key:
            scenario_variations[key]=[]
        for value in column[1:]:
            if value is not None:
                scenario_variations[key].append(value)

    return simulation_config, model_parameters, timeseries_data, scenario_variations


def count_simulations(scenario_variations):
    """
    This is just a small helper function that calculates the total number of simulation runs required to calculate all possible combinations of the given input parameter variations.

    :param scenario_variations: [dataframe] containing a list of possible variations for all variable simulation parameters as specified in DRI_Parameters.xlsx\"3 Scenario Variations"
    :return: [int]
    """
    return len(scenario_variations["avg_electricity_prices"]) * \
           len(scenario_variations["market_timeseries"]) * \
           len(scenario_variations["natural_gas_cost"]) * \
           len(scenario_variations["volatilities"]) * \
           len(scenario_variations["CO2_prices"]) * \
           len(scenario_variations["plant_types"]) * \
           len(scenario_variations["CO2_reduction"]) * \
           len(scenario_variations["electricity_emissions"]) * \
        12 # month


def create_simulation_list(simulation_config, model_parameters, timeseries_data, scenario_variations):

    # initialize empty list
    simulation_list = list()

    # iterate over all specified scenario variations
    for plant_type in scenario_variations["plant_types"]:
        # create the basic plant layout and a static base-scenario
        [factory, scenario] = dri.create_steel_plant(model_parameters,
                                                     plant_type,
                                                     config={"enable_log": simulation_config["enable_log_factory_setup"],
                                                             "enable_slacks": simulation_config["enable_slacks"]})

        for avg_electricity_price in scenario_variations["avg_electricity_prices"]:
            for natural_gas_cost in scenario_variations["natural_gas_cost"]:
                for volatility in scenario_variations["volatilities"]:
                    for co2_price in scenario_variations["CO2_prices"]:
                        for co2_reduction in scenario_variations["CO2_reduction"]:
                            for timeseries in scenario_variations["market_timeseries"]:
                                for electricity_emissions in scenario_variations["electricity_emissions"]:
                                    for month in list(range(1, 13)):
                                        simulation_list.append({"avg_electricity_price": avg_electricity_price,
                                                                "layout": plant_type,
                                                                "natural_gas_cost": natural_gas_cost,
                                                                "volatility": volatility,
                                                                "co2_price": co2_price,
                                                                "co2_reduction": co2_reduction,
                                                                "scenario": scenario,
                                                                "factory": factory,
                                                                "month": month,
                                                                "timeseries": timeseries,
                                                                "cost_electricity": timeseries_data[f"{timeseries}_cost"][
                                                                                    (month - 1) * factory.timesteps + 1:month * factory.timesteps+1],
                                                                "emissions_electricity": timeseries_data[
                                                                                             f"{timeseries}_emissions"][
                                                                                         (month - 1) * factory.timesteps + 1:month * factory.timesteps+1]*electricity_emissions,
                                                                "electricity_emissions": electricity_emissions,
                                                                "simulation_number": len(simulation_list) + 1})
    return simulation_list