# IMPORTS
import logging

import numpy as np
import openpyxl
import pandas as pd
import xlsxwriter

import factory_flexibility_model.factory.Blueprint as bp
import factory_flexibility_model.simulation.Simulation as sim

# MAIN FUNCTIONS


def write_results_to_xlsx(
    simulation: sim.Simulation, path: str, *, filename: str = None
):
    """
    This function checks, if the energy/material - conservation at a Component has been fulfilled during Simulation
    :param simulation: [Simulation.Simulation-object]
    :param path: [string] path where the created excel-file shall be stored
    :param name: [string] Optional: Name of the excel file, if left empty, the file will be called Simulation.name_results.xlsx
    :return: [Boolean] True if saving was successfull
    """

    # check, that the Simulation has already been performed
    if not simulation.simulated:
        logging.warning(
            "The Simulation has to be solved before the results can be exported. Calling the Simulation method now..."
        )
        simulation.simulate()

    # check, that results have been collected
    if not simulation.results_collected:
        logging.warning(
            "The Simulation results have not been processed yet. Calling the result processing method now..."
        )
        simulation.__collect_results()

    # create a result dict with reduced depth:
    result_dict = {}
    for key, value in simulation.result.items():
        if isinstance(value, dict):
            for subkey, subvalue in value.items():
                result_dict[f"{key}_{subkey}"] = subvalue
        else:
            result_dict[key] = value

    # create a new excel workbook
    if filename is None:
        workbook = xlsxwriter.Workbook(f"{path}\\{simulation.name}_results.xlsx")
    else:
        workbook = xlsxwriter.Workbook(f"{path}\\{filename}.xlsx")

    worksheet = workbook.add_worksheet()

    # iterate over all parameters and write each of them in an individual column in the excel file
    col = 0
    for key in result_dict.keys():
        # insert key names
        worksheet.write(0, col, key)

        # insert data
        if (
            isinstance(result_dict[key], int)
            or isinstance(result_dict[key], float)
            or isinstance(result_dict[key], bool)
        ):
            worksheet.write(1, col, result_dict[key])
        else:
            for row in range(len(result_dict[key])):
                worksheet.write(row + 1, col, result_dict[key][row])

        # continue with next column in excel sheet
        col += 1

    workbook.close()


def import_xlsx_to_blueprint(data_path: str):
    """
    This function takes an Excel-File in the factory-template format and returns a blueprint-object
    containing all the specified components and configurations in the excel-file.
    :param data_path: [string] Filepath of the Excel file to import from
    :return: [factory.blueprint]
    """

    # OPEN EXCEL-FILE
    workbook = openpyxl.load_workbook(data_path)

    # CREATE FACTORY
    blueprint = __create_blueprint(workbook)

    # CREATE FLOWS
    blueprint = __add_flows_to_blueprint(workbook, blueprint)

    # READ IN SCENARIO REQUIREMENTS
    scenario_parameters = __read_in_scenario_parameters(workbook["SCENARIO PARAMETERS"])

    # READ IN TIMESERIES AND SCHEDULER DEMANDS
    timeseries = __read_in_timeseries(data_path)
    # scheduler_demands = __read_in_scheduler_demands(workbook["SCHEDULER DEMANDS"])

    # CREATE COMPONENTS
    blueprint = __add_components_to_blueprint(
        workbook, blueprint, timeseries, scenario_parameters
    )

    # CREATE CONNECTIONS
    blueprint = __add_connections_to_blueprint(workbook, blueprint)

    return blueprint


def import_xlsx_to_factory(data_path: str):
    """
    This function takes an Excel-File in the factory-template format and returns a factory-object
    containing all the specified components and configurations in the excel-file. This is done by first transforming
    the excel information into a factory-blueprint-object and then calling the import_factory_blueprint-method to create
    the factory object itself
    :param data_path: [string] Filepath of the Excel file to import from
    :return: [factory.factory]
    """

    # CREATE THE BLUEPRINT
    blueprint = import_xlsx_to_blueprint(data_path)

    # CREATE THE FACTORY OBJECT
    factory = blueprint.to_factory()

    return factory


# UTILITY FUNCTIONS


def __add_flows_to_blueprint(workbook: openpyxl.workbook, blueprint: bp.blueprint):
    """
    This function adds all flows specified in the given workbook
    to the given factory and then returns the edited factory blueprint
    :param workbook: excel-workbook in openpyxl-format
    :param factory:  blueprint.blueprint-object
    :return: blueprint.blueprint-object
    """

    # read in the required sheet
    input_data_flows = __read_component_sheet(workbook["FLOWS"])

    # iterate over all given flows
    for flow_name in input_data_flows:
        # get data corresponding to the flowtype name
        flow_data = input_data_flows[flow_name]

        # initialize new flowtype within blueprint
        blueprint.flows[flow_name] = {}

        # add given specifications to the flowtype description within the blueprint
        blueprint.flows[flow_name]["name"] = flow_name
        for attribute in flow_data:
            blueprint.flows[flow_name][attribute] = flow_data[attribute]

    return blueprint


def __add_components_to_blueprint(
    workbook: openpyxl.workbook,
    blueprint: bp.blueprint,
    timeseries: str,
    scenario_parameters: str,
):
    """
    This function adds all components specified in the given workbook to the given
    factory blueprint and then returns the edited blueprint
    :param workbook: excel-workbook in openpyxl-format
    :param factory:  blueprint-object
    :return: blueprint-object
    """
    types = {
        "pool": "POOLS",
        "sink": "SINKS",
        "converter": "CONVERTERS",
        "source": "SOURCES",
        "deadtime": "DEADTIMES",
        "storage": "STORAGES",
        "schedule": "SCHEDULERS",
        "thermalsystem": "THERMALSYSTEMS",
        "triggerdemand": "TRIGGERDEMANDS",
    }
    for component_type in types:
        input_data_component = __read_component_sheet(workbook[types[component_type]])
        # iterate over all given components
        for component_name in input_data_component:
            component_data = input_data_component[component_name]

            # create a list of specifications for the Component
            kwargs = __create_component_attribute_list(
                component_data, timeseries, scenario_parameters
            )

            blueprint.components[component_name] = kwargs
            blueprint.components[component_name]["name"] = component_name
            blueprint.components[component_name]["type"] = component_type

    return blueprint


def __add_connections_to_blueprint(
    workbook: openpyxl.workbook, blueprint: bp.blueprint
):
    """
    This function adds all connections specified in the given workbook
    to the given factory blueprint and then returns the edited blueprint
    :param workbook: excel-workbook in openpyxl-format
    :param factory:  blueprint-object
    :return: blueprint-object
    """

    # initialize variables
    parameter_names = []

    # define the relevant workbook sheet
    sheet = workbook["CONNECTIONS"]

    # iterate over all rows
    for row in sheet.rows:
        # Are we looking at the first row?
        if row[1].value:
            if row[0].value == "name":
                # Yes: generate a list of the existing parameters
                parameter_names = ["name"]
                for cell in row[1:]:
                    if cell.value:
                        parameter_names.append(cell.value)
            else:
                # no: collect all data about the current connection
                connection_parameters = {}
                for cell in row[1:]:
                    if cell.value:
                        connection_parameters[
                            parameter_names[cell.column - 1]
                        ] = cell.value

                # create list of parameters and values
                kwargs = {}
                for attribute in connection_parameters:
                    kwargs[attribute] = connection_parameters[attribute]
                # add the connection to the blueprint with given parameters
                blueprint.connections[f"{kwargs['from']}_to_{kwargs['to']}"] = kwargs

    return blueprint


def __create_blueprint(workbook: openpyxl.workbook):
    """
    This function takes an Excel-workbook and creates a new factory.blueprint-object with the
    specifications given in the "GENERAL"-Sheet
    :param workbook: [openpyxl-workbook] Excel-workbook in openpyxl-format
    :return: [factory.blueprint]
    """

    # Read in excel-sheet data
    sheet = workbook["GENERAL"]

    # iterate over all rows to collect the sheet data in a dictionary
    sheet_data = {}
    for row in sheet.rows:
        # if row contains data: add it to the dictionary
        if row[0].value:
            sheet_data[row[0].value] = row[1].value

    # create a newe blueprint with the given parameters
    blueprint = bp.blueprint()
    blueprint.info["name"] = sheet_data["factory_name"]
    blueprint.info["description"] = sheet_data["factory_description"]
    blueprint.info["max_timesteps"] = sheet_data["max_number_of_timesteps"]
    blueprint.info["enable_slacks"] = sheet_data["enable_slacks"]

    return blueprint


def __create_component_attribute_list(component_data, timeseries, scenario_parameters):
    """
    This function takes the configuration sheet of one Component as well as the list of all timeseries
    and returns a dictionary with all the information needed to create and configure the Component.
    :param component_data: openpyxl-object with the sheet_data of the Component
    :param timeseries: pandas-dataframe of the TIMESERIES-Sheet of a configuration-excel-document
    :return: kwargs (dict)
    """

    # create a list of the given specifications
    kwargs = {}

    # iterate over all columns of the Component-specific table
    for attribute in component_data:
        value = component_data[attribute]

        # check, if the given value is a key to a timeseries
        if value in timeseries:
            # if yes: add the timeseries instead of the given value
            kwargs[attribute] = timeseries[value]

        elif value in scenario_parameters:
            # if yes: add the timeseries instead of the given value
            kwargs["scenario_data"] = {attribute: value}
        else:
            # if no: add the given value to the specification list
            kwargs[attribute] = value

    return kwargs


def __read_component_sheet(sheet):
    """
    This function takes an Excel sheet in openpyxl-format. It should be one
    of the Component-specific sheets of the Factory-Template.xlsx-file
    It returns the information on all the specified components as a dictionary.
    :param sheet: openpyxl-workbook-sheet (Factor_Template.xlsx-Format)
    :return:      dictionary
    """

    # initialize variables
    sheet_data = {}
    parameter_names = []

    # Iteration over all rows in the sheet
    for row in sheet.rows:

        # is there data in the current row?
        if not row[0].value:
            continue

        # extract the key of the row
        key = row[0].value

        # if current row contains the headers: create list of attributes
        if key == "name":
            parameter_names = ["name"]
            for cell in row[1:]:
                if cell.value:
                    parameter_names.append(cell.value)

        # otherwise: extract all the specifications for available attributes
        else:

            # create empty dictionary
            sheet_data[row[0].value] = {}

            # iterate over all columns/attributes
            for cell in row[1:]:

                # check if there is a value specified
                if cell.value:

                    # if yes: add it to the list of parameters
                    sheet_data[row[0].value][
                        parameter_names[cell.column - 1]
                    ] = cell.value
    return sheet_data


def __read_in_scenario_parameters(sheet):
    """
    This function takes the "SCENARIO PARAMETERS" - workbook sheet of a factory
    template xlsx and returns alist of all the keys
    :param sheet: "SCENARIO PARAMETERS" sheet from excel validate data in openpyxl format
    :return: List containing all specified keys from the given excel-file
    """

    # initialize an empty list
    scenario_parameters = []

    # iterate over all rows
    for row in sheet.rows:

        # abort if there is no value in the current row
        if not row[0].value:
            continue

        # abort if the current row contains the header
        if row[0].value == "Required Parameters in the scenario:":
            continue

        # add the current value to the list of scenario parameters
        scenario_parameters.append(row[0].value)

    # return the list
    return scenario_parameters


def __read_in_scheduler_demands(sheet):
    """
    This function takes a "SCHEDULER DEMANDS"-workbook sheet of a factory template.xlsx and returns a dictionary with
    all the contained scheduler demands lists as numpy matrices.
    :param sheet: "SCHEDULER DEMANDS" sheet from factory_template.xlsx in openpyxl format
    :return: dictionary with all specified part demands
    """

    # initialize an empty dict
    scheduler_demands = {}

    # initialize a variable to store the name of the currently handled demand
    current_demand_name = ""

    # iterate over all columns
    for col in sheet.columns:

        # check, if a new demand starts at the current column:
        if col[0].value is not None:

            # if yes: set the current demands name to the key
            current_demand_name = col[0].value

            # create a new item within scheduler_demands
            scheduler_demands[current_demand_name] = []

        # create an array with all given values of the row
        row_values = []
        for cell in col[2:]:
            if cell.value is not None:
                row_values.append(cell.value)
        row_values = np.array(row_values)

        # if there were any values: write the timeseries to the scheduler_demands-dict
        if not row_values == []:
            scheduler_demands[current_demand_name] = np.concatenate(
                (scheduler_demands[current_demand_name], row_values)
            )

    # check, if the format is correct
    for demand in scheduler_demands.values():
        if not len(demand) == 4:
            logging.critical(
                "At least one partdemand of sheet 'SCHEDULER DEMANDS' is given in an invalid way!"
            )
            raise Exception

    # return the list
    return scheduler_demands


def __read_in_timeseries(data_path: str):
    """
    This function takes the path to an excel-file and returns all the timeseries from
    the "TIMESERIES"-Sheet as a Pandas frame
    :param data_path: [string] Path to the user-given Excel file
    :return: [pd.dataframe] containing all specified timeseries from the given excel-file
    """

    timeseries = pd.read_excel(data_path, sheet_name="TIMESERIES")
    return timeseries
